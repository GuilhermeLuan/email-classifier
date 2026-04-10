import random
import string
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import copy

# ── Geração de CPF válido ───────────────────────────────────────────────────

def calcular_digito(cpf_parcial):
    n = len(cpf_parcial) + 1
    soma = sum(int(d) * (n - i) for i, d in enumerate(cpf_parcial))
    resto = soma % 11
    return 0 if resto < 2 else 11 - resto

def gerar_cpf():
    while True:
        base = [random.randint(0, 9) for _ in range(9)]
        d1 = calcular_digito(base)
        d2 = calcular_digito(base + [d1])
        cpf = base + [d1, d2]
        # Rejeita CPFs com todos os dígitos iguais
        if len(set(cpf)) > 1:
            return int("".join(map(str, cpf)))

# ── Geração de matrícula única (9 dígitos) ──────────────────────────────────

def gerar_matricula():
    return random.randint(100_000_000, 999_999_999)

# ── Nomes aleatórios brasileiros ────────────────────────────────────────────

PRIMEIROS = [
    "Ana", "Carlos", "Maria", "João", "Fernanda", "Lucas", "Beatriz", "Rafael",
    "Juliana", "Pedro", "Camila", "Marcos", "Larissa", "André", "Patrícia",
    "Felipe", "Gabriela", "Rodrigo", "Mariana", "Thiago", "Amanda", "Bruno",
    "Letícia", "Diego", "Natália", "Gustavo", "Aline", "Henrique", "Renata",
    "Leonardo", "Vanessa", "Eduardo", "Sandra", "Vinicius", "Priscila",
    "Mateus", "Cristina", "Alexandre", "Tatiana", "Fábio", "Érica",
    "Leandro", "Simone", "Ricardo", "Débora", "Sérgio", "Alessandra",
    "Daniel", "Mônica", "Adriano",
]

SOBRENOMES = [
    "Silva", "Santos", "Oliveira", "Souza", "Rodrigues", "Ferreira", "Alves",
    "Pereira", "Lima", "Gomes", "Costa", "Ribeiro", "Martins", "Carvalho",
    "Almeida", "Lopes", "Sousa", "Fernandes", "Vieira", "Barbosa", "Rocha",
    "Dias", "Nascimento", "Andrade", "Moreira", "Nunes", "Marques", "Machado",
    "Mendes", "Freitas", "Cardoso", "Ramos", "Araújo", "Melo", "Cavalcanti",
    "Correia", "Teixeira", "Cunha", "Pinto", "Azevedo", "Monteiro", "Borges",
    "Medeiros", "Moraes", "Castro", "Miranda", "Neto", "Fonseca", "Pires",
]

def gerar_nome(usados):
    tentativas = 0
    while tentativas < 1000:
        nome = f"{random.choice(PRIMEIROS)} {random.choice(SOBRENOMES)} {random.choice(SOBRENOMES)}"
        if nome not in usados:
            usados.add(nome)
            return nome
        tentativas += 1
    # Fallback com sufixo numérico
    base = f"{random.choice(PRIMEIROS)} {random.choice(SOBRENOMES)}"
    sufixo = random.randint(100, 9999)
    nome = f"{base} {sufixo}"
    usados.add(nome)
    return nome

# ── Dados fixos conforme modelo ─────────────────────────────────────────────

SEXO            = "M - Masculino"
ETNIA           = "1 - Branca"
GRAU_INSTRUCAO  = "01 - Analfabeto, inclusive o que, embora tenha recebido instrução, não se alfabetizou"
DATA_NASCIMENTO = datetime(1990, 1, 1)
PAIS_NASCIMENTO = "105 - Brasil"
PAIS_NACIOANAL  = "105 - Brasil"
CEP             = 72005607
LOGRADOURO      = "Logradouro teste"
NUMERO          = 125
UF              = "AC"
MUNICIPIO       = "1200013 - Acrelândia"
TIPO_CADASTRO   = "N - Não (Início de TSVE)"
DATA_INICIO     = datetime(2024, 2, 1)
NOME_CARGO      = "Médico"
CBO             = 225125

HEADERS = [
    "cpf", "nome", "sexo", "etnia", "grau_instrucao", "data_nascimento",
    "pais_nascimento", "pais_nacionalidade", "cep", "logradouro", "numero",
    "uf", "municipio", "tipo_cadastro", "matricula", "data_inicio",
    "nome_cargo", "cbo",
]

COL_WIDTHS = {
    "A": 12.0,
    "B": 12.85546875,
    "C": 10.28515625,
    "D": 9.85546875,
    "E": 78.0,
    "F": 16.42578125,
    "G": 16.28515625,
    "H": 18.85546875,
    "I": 8.43,
    "J": 10.85546875,
    "K": 8.43,
    "L": 10.28515625,
    "M": 10.0,
    "N": 28.85546875,
    "O": 10.0,
    "P": 10.42578125,
    "Q": 11.85546875,
    "R": 8.43,
}

# ── Geração da planilha ──────────────────────────────────────────────────────

def gerar_planilha(qtd: int, caminho_saida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    # Cabeçalho
    ws.append(HEADERS)
    font_header = Font(name="Aptos Narrow")
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=col).font = font_header

    # Gerar registros únicos
    cpfs_usados      = set()
    matriculas_usadas = set()
    nomes_usados      = set()

    for _ in range(qtd):
        # CPF único e válido
        while True:
            cpf = gerar_cpf()
            if cpf not in cpfs_usados:
                cpfs_usados.add(cpf)
                break

        # Matrícula única de 9 dígitos
        while True:
            mat = gerar_matricula()
            if mat not in matriculas_usadas:
                matriculas_usadas.add(mat)
                break

        nome = gerar_nome(nomes_usados)

        row = [
            cpf,
            nome,
            SEXO,
            ETNIA,
            GRAU_INSTRUCAO,
            DATA_NASCIMENTO,
            PAIS_NASCIMENTO,
            PAIS_NACIOANAL,
            CEP,
            LOGRADOURO,
            NUMERO,
            UF,
            MUNICIPIO,
            TIPO_CADASTRO,
            mat,
            DATA_INICIO,
            NOME_CARGO,
            CBO,
        ]
        ws.append(row)

    # Formatar datas (colunas F=6 e P=16)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [6, 16]:
            cell = row[col_idx - 1]
            cell.number_format = "mm-dd-yy"

    # Larguras de colunas
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(caminho_saida)
    print(f"✅ Planilha gerada com {qtd} registros em: {caminho_saida}")


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    while True:
        try:
            qtd = int(input("Quantos registros deseja gerar? ").strip())
            if qtd <= 0:
                print("Digite um número maior que zero.")
                continue
            break
        except ValueError:
            print("Entrada inválida. Digite um número inteiro.")

    saida = f"s2300_{qtd}_registros.xlsx"
    gerar_planilha(qtd, saida)